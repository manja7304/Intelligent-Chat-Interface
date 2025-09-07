"""
AI-powered form filler for generating HR forms based on candidate data
"""

import json
import logging
from typing import Dict, List, Optional, Any
import pandas as pd
from datetime import datetime
import os
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

logger = logging.getLogger(__name__)


DEFAULT_MODEL = "gpt-3.5-turbo"
DEFAULT_MAX_TOKENS = 2000
DEFAULT_TEMPERATURE = 0.7
OPENAI_CHAT_COMPLETIONS_URL = "https://api.openai.com/v1/chat/completions"
DEFAULT_TIMEOUT_SECONDS = 60


class AIFormFiller:
    """Uses AI to intelligently fill HR forms based on candidate data"""

    def __init__(
        self,
        api_key: str,
        *,
        model: str = DEFAULT_MODEL,
        max_tokens: int = DEFAULT_MAX_TOKENS,
        temperature: float = DEFAULT_TEMPERATURE,
        timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS,
        session: Optional[requests.Session] = None,
    ):
        self.api_key = api_key
        self.model = model
        self.max_tokens = max_tokens
        self.temperature = temperature
        self.timeout_seconds = timeout_seconds

        # Create or use provided HTTP session with retries
        self.session = session or self._build_session_with_retries()

    def generate_hr_form(
        self, candidate_data: Dict[str, Any], form_template: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Generate filled HR form based on candidate data and form template

        Args:
            candidate_data: Structured candidate information
            form_template: HR form template with fields and instructions

        Returns:
            Filled HR form data
        """
        try:
            # Check if API key is available
            if not self.api_key or self.api_key == "your_openai_api_key_here":
                logger.warning("OpenAI API key not provided, using fallback form")
                return self._create_fallback_form(form_template, candidate_data)

            # Prepare the prompt for AI
            prompt = self._create_form_filling_prompt(candidate_data, form_template)

            # Call OpenAI API via REST to avoid SDK init issues
            ai_response = self._call_openai_chat(prompt)

            # Parse the AI response
            filled_form = self._parse_ai_response(
                ai_response, form_template, candidate_data
            )
            # Normalize to expected structure
            filled_form = self._normalize_filled_form(
                filled_form, form_template, candidate_data
            )

            # Add metadata
            filled_form["_metadata"] = {
                "generated_at": datetime.now().isoformat(),
                "candidate_id": candidate_data.get("id"),
                "form_type": form_template.get("form_type", "standard"),
                "ai_model": self.model,
            }

            logger.info(
                f"Generated HR form for candidate: {candidate_data.get('name', 'Unknown')}"
            )
            return filled_form

        except Exception as e:
            logger.error(f"Error generating HR form: {e}")
            raise

    def _build_session_with_retries(self) -> requests.Session:
        """Create a requests session configured with retries and backoff."""
        session = requests.Session()
        retry = Retry(
            total=3,
            read=3,
            connect=3,
            backoff_factor=0.5,
            status_forcelist=(429, 500, 502, 503, 504),
            allowed_methods=("POST", "GET"),
            raise_on_status=False,
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        return session

    def _create_form_filling_prompt(
        self, candidate_data: Dict[str, Any], form_template: Dict[str, Any]
    ) -> str:
        """Create a comprehensive prompt for AI form filling"""

        prompt = f"""
        Please fill out the following HR form based on the provided candidate data.
        
        CANDIDATE DATA:
        {json.dumps(candidate_data, indent=2)}
        
        FORM TEMPLATE:
        {json.dumps(form_template, indent=2)}
        
        INSTRUCTIONS:
        1. Fill in all applicable fields based on the candidate data
        2. If information is not available, leave the field empty or mark as "Not Available"
        3. Use professional language and formatting
        4. Ensure all dates are in YYYY-MM-DD format
        5. For skills, list them in order of relevance/importance
        6. For experience, include key achievements and responsibilities
        7. Maintain consistency with the original candidate data
        
        Please return the filled form as a JSON object with the same structure as the template.
        """

        return prompt

    def _parse_ai_response(
        self,
        ai_response: str,
        form_template: Dict[str, Any],
        candidate_data: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Parse AI response and structure it according to the form template"""
        try:
            # Prefer fenced JSON blocks if present
            fenced_start = ai_response.find("```")
            if fenced_start != -1:
                fenced_end = ai_response.find("```", fenced_start + 3)
                if fenced_end != -1:
                    fenced_content = ai_response[fenced_start + 3 : fenced_end].strip()
                    # Remove potential language hints like ```json
                    if fenced_content.lower().startswith("json\n"):
                        fenced_content = fenced_content[5:]
                    return json.loads(fenced_content)

            # Fallback: extract first JSON object heuristically
            json_start = ai_response.find("{")
            json_end = ai_response.rfind("}") + 1
            if json_start != -1 and json_end > json_start:
                return json.loads(ai_response[json_start:json_end])

            # If nothing parseable, create structured fallback
            return self._create_fallback_form(form_template, candidate_data)

        except json.JSONDecodeError:
            logger.warning("Failed to parse AI response as JSON, using fallback")
            return self._create_fallback_form(form_template, candidate_data)

    def _create_fallback_form(
        self, form_template: Dict[str, Any], candidate_data: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """Create a fallback form when AI response parsing fails"""
        filled_form = {}

        for section, fields in form_template.get("sections", {}).items():
            filled_form[section] = {}
            for field, config in fields.items():
                # Try to map candidate data to form fields
                if candidate_data:
                    field_value = self._map_candidate_data_to_field(
                        field, candidate_data
                    )
                    if field_value:
                        filled_form[section][field] = field_value
                    else:
                        filled_form[section][field] = config.get("default_value", "")
                else:
                    filled_form[section][field] = config.get("default_value", "")

        return filled_form

    def _normalize_filled_form(
        self,
        filled_form: Dict[str, Any],
        form_template: Dict[str, Any],
        candidate_data: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Ensure the returned form matches the template section/field structure."""
        try:
            template_sections = form_template.get("sections", {})

            # If model returned {"sections": {...}}, unwrap it
            if (
                isinstance(filled_form, dict)
                and "sections" in filled_form
                and isinstance(filled_form["sections"], dict)
            ):
                filled_form = filled_form["sections"]

            if not isinstance(filled_form, dict):
                return self._create_fallback_form(form_template, candidate_data)

            normalized: Dict[str, Any] = {}
            for section_name, fields in template_sections.items():
                # Accept case-insensitive matches for section names
                raw_section = {}
                for key, val in filled_form.items():
                    if isinstance(key, str) and key.lower() == section_name.lower():
                        raw_section = val
                        break
                if not raw_section:
                    raw_section = filled_form.get(section_name, {})
                if not isinstance(raw_section, dict):
                    raw_section = {"text": raw_section} if raw_section else {}

                ensured_section: Dict[str, Any] = {}
                for field_key, field_cfg in fields.items():
                    # Case-insensitive field matching
                    val = None
                    for rk, rv in raw_section.items():
                        if isinstance(rk, str) and rk.lower() == field_key.lower():
                            val = rv
                            break
                    if val is None:
                        val = raw_section.get(field_key)
                    if isinstance(val, (str, int, float)) and val is not None:
                        ensured_section[field_key] = val
                    else:
                        mapped = self._map_candidate_data_to_field(
                            field_key, candidate_data
                        )
                        ensured_section[field_key] = mapped or field_cfg.get(
                            "default_value", ""
                        )

                # Keep extra simple fields from AI, without overwriting expected ones
                for extra_key, extra_val in raw_section.items():
                    if extra_key not in ensured_section and isinstance(
                        extra_val, (str, int, float)
                    ):
                        ensured_section[extra_key] = extra_val

                normalized[section_name] = ensured_section

            return normalized
        except Exception:
            return self._create_fallback_form(form_template, candidate_data)

    def _map_candidate_data_to_field(
        self, field_name: str, candidate_data: Dict[str, Any]
    ) -> str:
        """Map candidate data to form field names"""

        # Helper function to safely parse JSON strings
        def safe_parse_json(data, default=None):
            if default is None:
                default = []
            if isinstance(data, str):
                try:
                    import json

                    return json.loads(data)
                except (json.JSONDecodeError, TypeError):
                    return default
            return data if data is not None else default

        # Get experience data (handle both list and string formats)
        experience_data = candidate_data.get("experience", [])
        if isinstance(experience_data, str):
            experience_data = safe_parse_json(experience_data, [])

        # Get education data (handle both list and string formats)
        education_data = candidate_data.get("education", [])
        if isinstance(education_data, str):
            education_data = safe_parse_json(education_data, [])

        # Get skills data (handle both list and string formats)
        skills_data = candidate_data.get("skills", [])
        if isinstance(skills_data, str):
            skills_data = safe_parse_json(skills_data, [])

        field_mapping = {
            "full_name": candidate_data.get("name", ""),
            "email": candidate_data.get("email", ""),
            "phone": candidate_data.get("phone", ""),
            "location": candidate_data.get("location", ""),
            "linkedin_url": candidate_data.get("linkedin_url", ""),
            "current_position": candidate_data.get("current_position", ""),
            "current_company": candidate_data.get("current_company", ""),
            "experience_years": candidate_data.get("experience_years", ""),
            "technical_skills": ", ".join(skills_data) if skills_data else "",
            "work_experience": (
                ", ".join(
                    [
                        (
                            (exp.get("title", "") + " at " + exp.get("company", ""))
                            if isinstance(exp, dict)
                            else str(exp)
                        )
                        for exp in experience_data
                    ]
                )
                if experience_data
                else ""
            ),
            "education": (
                ", ".join(
                    [
                        (
                            (
                                edu.get("degree", "")
                                + " from "
                                + edu.get("institution", "")
                            )
                            if isinstance(edu, dict)
                            else str(edu)
                        )
                        for edu in education_data
                    ]
                )
                if education_data
                else ""
            ),
        }

        return field_mapping.get(field_name, "")

    def generate_standard_hr_form(
        self, candidate_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Generate a standard HR form for the candidate"""

        standard_template = {
            "form_type": "standard_hr_form",
            "sections": {
                "personal_information": {
                    "full_name": {
                        "label": "Full Name",
                        "type": "text",
                        "required": True,
                    },
                    "email": {
                        "label": "Email Address",
                        "type": "email",
                        "required": True,
                    },
                    "phone": {"label": "Phone Number", "type": "tel", "required": True},
                    "location": {
                        "label": "Location",
                        "type": "text",
                        "required": False,
                    },
                    "linkedin_url": {
                        "label": "LinkedIn Profile",
                        "type": "url",
                        "required": False,
                    },
                },
                "professional_summary": {
                    "summary": {
                        "label": "Professional Summary",
                        "type": "textarea",
                        "required": True,
                    },
                    "current_position": {
                        "label": "Current Position",
                        "type": "text",
                        "required": True,
                    },
                    "current_company": {
                        "label": "Current Company",
                        "type": "text",
                        "required": True,
                    },
                    "experience_years": {
                        "label": "Years of Experience",
                        "type": "number",
                        "required": True,
                    },
                },
                "skills_assessment": {
                    "technical_skills": {
                        "label": "Technical Skills",
                        "type": "textarea",
                        "required": True,
                    },
                    "soft_skills": {
                        "label": "Soft Skills",
                        "type": "textarea",
                        "required": False,
                    },
                    "certifications": {
                        "label": "Certifications",
                        "type": "textarea",
                        "required": False,
                    },
                },
                "experience_details": {
                    "work_experience": {
                        "label": "Work Experience",
                        "type": "textarea",
                        "required": True,
                    },
                    "key_achievements": {
                        "label": "Key Achievements",
                        "type": "textarea",
                        "required": False,
                    },
                },
                "education_background": {
                    "education": {
                        "label": "Education",
                        "type": "textarea",
                        "required": True,
                    },
                    "additional_training": {
                        "label": "Additional Training/Courses",
                        "type": "textarea",
                        "required": False,
                    },
                },
                "hr_assessment": {
                    "availability": {
                        "label": "Availability",
                        "type": "text",
                        "required": False,
                    },
                    "salary_expectations": {
                        "label": "Salary Expectations",
                        "type": "text",
                        "required": False,
                    },
                    "notice_period": {
                        "label": "Notice Period",
                        "type": "text",
                        "required": False,
                    },
                    "additional_notes": {
                        "label": "Additional Notes",
                        "type": "textarea",
                        "required": False,
                    },
                },
            },
        }

        return self.generate_hr_form(candidate_data, standard_template)

    def generate_interview_form(self, candidate_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate an interview assessment form"""

        interview_template = {
            "form_type": "interview_assessment",
            "sections": {
                "candidate_overview": {
                    "candidate_name": {
                        "label": "Candidate Name",
                        "type": "text",
                        "required": True,
                    },
                    "position_applied": {
                        "label": "Position Applied For",
                        "type": "text",
                        "required": True,
                    },
                    "interview_date": {
                        "label": "Interview Date",
                        "type": "date",
                        "required": True,
                    },
                    "interviewer_name": {
                        "label": "Interviewer Name",
                        "type": "text",
                        "required": True,
                    },
                },
                "technical_assessment": {
                    "technical_skills_rating": {
                        "label": "Technical Skills Rating (1-5)",
                        "type": "number",
                        "required": True,
                    },
                    "problem_solving_ability": {
                        "label": "Problem Solving Ability (1-5)",
                        "type": "number",
                        "required": True,
                    },
                    "technical_notes": {
                        "label": "Technical Assessment Notes",
                        "type": "textarea",
                        "required": False,
                    },
                },
                "communication_assessment": {
                    "communication_skills": {
                        "label": "Communication Skills (1-5)",
                        "type": "number",
                        "required": True,
                    },
                    "presentation_ability": {
                        "label": "Presentation Ability (1-5)",
                        "type": "number",
                        "required": True,
                    },
                    "communication_notes": {
                        "label": "Communication Assessment Notes",
                        "type": "textarea",
                        "required": False,
                    },
                },
                "cultural_fit": {
                    "team_work": {
                        "label": "Team Work (1-5)",
                        "type": "number",
                        "required": True,
                    },
                    "adaptability": {
                        "label": "Adaptability (1-5)",
                        "type": "number",
                        "required": True,
                    },
                    "cultural_fit_notes": {
                        "label": "Cultural Fit Notes",
                        "type": "textarea",
                        "required": False,
                    },
                },
                "overall_assessment": {
                    "overall_rating": {
                        "label": "Overall Rating (1-5)",
                        "type": "number",
                        "required": True,
                    },
                    "recommendation": {
                        "label": "Recommendation",
                        "type": "select",
                        "options": ["Strong Hire", "Hire", "No Hire", "Strong No Hire"],
                        "required": True,
                    },
                    "strengths": {
                        "label": "Key Strengths",
                        "type": "textarea",
                        "required": True,
                    },
                    "areas_for_improvement": {
                        "label": "Areas for Improvement",
                        "type": "textarea",
                        "required": False,
                    },
                    "final_notes": {
                        "label": "Final Notes",
                        "type": "textarea",
                        "required": False,
                    },
                },
            },
        }

        return self.generate_hr_form(candidate_data, interview_template)

    def export_form_to_pdf(self, filled_form: Dict[str, Any], output_path: str) -> str:
        """Export filled form to PDF using reportlab"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch

            doc = SimpleDocTemplate(output_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=16,
                spaceAfter=30,
                alignment=1,  # Center alignment
            )

            form_type = filled_form.get("_metadata", {}).get("form_type", "HR Form")
            title = Paragraph(f"{form_type.upper()}", title_style)
            story.append(title)
            story.append(Spacer(1, 20))

            # Add form sections
            for section_name, section_data in filled_form.items():
                if section_name.startswith("_"):
                    continue

                # Section header
                section_style = ParagraphStyle(
                    "SectionHeader",
                    parent=styles["Heading2"],
                    fontSize=14,
                    spaceAfter=12,
                    spaceBefore=20,
                )

                section_title = Paragraph(
                    section_name.replace("_", " ").title(), section_style
                )
                story.append(section_title)

                # Section content (handle both dict and string sections)
                if isinstance(section_data, dict):
                    for field_name, field_value in section_data.items():
                        if isinstance(field_value, (str, int, float)) and field_value:
                            field_text = f"<b>{field_name.replace('_', ' ').title()}:</b> {field_value}"
                            field_para = Paragraph(field_text, styles["Normal"])
                            story.append(field_para)
                            story.append(Spacer(1, 6))
                else:
                    # Render raw text for non-dict section content
                    if isinstance(section_data, (str, int, float)) and section_data:
                        field_para = Paragraph(str(section_data), styles["Normal"])
                        story.append(field_para)
                        story.append(Spacer(1, 6))

                story.append(Spacer(1, 12))

            doc.build(story)
            logger.info(f"Exported form to PDF: {output_path}")
            return output_path

        except ImportError:
            logger.error("reportlab not installed. Install with: pip install reportlab")
            raise
        except Exception as e:
            logger.error(f"Error exporting form to PDF: {e}")
            raise

    def export_form_to_excel(
        self, filled_form: Dict[str, Any], output_path: str
    ) -> str:
        """Export filled form to Excel using openpyxl"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "HR Form"

            # Style definitions
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

            row = 1

            # Add form sections
            for section_name, section_data in filled_form.items():
                if section_name.startswith("_"):
                    continue

                # Section header
                ws.cell(row=row, column=1, value=section_name.replace("_", " ").title())
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=1).fill = header_fill
                row += 1

                # Section fields (handle both dict and string sections)
                if isinstance(section_data, dict):
                    for field_name, field_value in section_data.items():
                        if isinstance(field_value, (str, int, float)) and field_value:
                            ws.cell(
                                row=row,
                                column=1,
                                value=field_name.replace("_", " ").title(),
                            )
                            ws.cell(row=row, column=2, value=str(field_value))
                            row += 1
                else:
                    if isinstance(section_data, (str, int, float)) and section_data:
                        ws.cell(row=row, column=1, value=str(section_data))
                        row += 1

                row += 1  # Empty row between sections

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(output_path)
            logger.info(f"Exported form to Excel: {output_path}")
            return output_path

        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise

    def _call_openai_chat(self, prompt: str) -> str:
        """Call OpenAI Chat Completions via REST and return the text response."""
        try:
            if not self.api_key or self.api_key == "your_openai_api_key_here":
                logger.warning("OpenAI API key not provided, using fallback form")
                return "{}"  # Return empty JSON to trigger fallback

            headers = {
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": self.model,
                "messages": [
                    {
                        "role": "system",
                        "content": "You are an expert HR assistant that fills out forms based on candidate data. Be accurate and professional.",
                    },
                    {"role": "user", "content": prompt},
                ],
                "max_tokens": self.max_tokens,
                "temperature": self.temperature,
            }
            resp = self.session.post(
                OPENAI_CHAT_COMPLETIONS_URL,
                headers=headers,
                json=payload,
                timeout=self.timeout_seconds,
            )
            resp.raise_for_status()
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            logger.error(f"Error calling OpenAI chat API: {e}")
            return "{}"  # Return empty JSON to trigger fallback
